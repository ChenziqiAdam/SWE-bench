"""Stage 1: Parse PRs.xlsx and fetch GitHub PR/issue content."""
from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Optional

import openpyxl

from swebench.collect.utils import Repo, PR_KEYWORDS
from swebench.eval_pipeline.constants import (
    COL_REPO, COL_PR_NUMBER, COL_TITLE,
    COL_CATEGORY, COL_ALGORITHM_NAME, COL_PAPER_REFERENCE,
    COL_HAS_TEST, COL_TEST_LINKS, COL_HAS_ISSUE, COL_URL,
)

logger = logging.getLogger(__name__)


def instance_ids_to_pr_filter(instance_ids: set[str]) -> dict[str, set[int]]:
    """
    Convert a set of instance_ids like {"scikit-learn__scikit-learn-31856"}
    into {repo: {pr_number}} so the ingest stage can skip unneeded rows.
    Instance ID format: owner__repo-pr_number  (double underscore, trailing dash+number)
    """
    result: dict[str, set[int]] = {}
    for iid in instance_ids:
        # Split on last "-" to get pr_number, rest is repo slug with __ separating owner/name
        repo_slug, pr_str = iid.rsplit("-", 1)
        repo_full = repo_slug.replace("__", "/", 1)
        result.setdefault(repo_full, set()).add(int(pr_str))
    return result


def normalize_issue_type(value) -> str:
    """Normalize spreadsheet issue/category type values for exact matching."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_spreadsheet(path: str, sheet: Optional[str] = None) -> list[dict]:
    """Read PRs.xlsx and return a list of row dicts."""
    wb = openpyxl.load_workbook(path)
    if sheet:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet!r} not found. Available: {wb.sheetnames}")
        ws = wb[sheet]
    else:
        ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, raw_row))
        # Skip rows with no repo or PR number
        if not row.get(COL_REPO) or not row.get(COL_PR_NUMBER):
            continue
        row[COL_PR_NUMBER] = int(row[COL_PR_NUMBER])
        rows.append(row)
    return rows


# Issues_v1.xlsx column names
_COL_ISSUE_NUMBER = "Issue Number"
_COL_CLOSING_PR = "Closing PR #"
_ISSUE_URL_RE = re.compile(r"github\.com/([^/\s]+/[^/\s]+)/issues/(\d+)")


def _issue_repo_from_url(url, fallback_repo: str) -> str:
    """The issue's own repo, parsed from its GitHub URL.

    The 'Repo' column always names the PR's repo. An issue can live in a
    different repo than the PR that closes it (e.g. a cross-repo "Fixes
    org/other-repo#N" reference) -- using the PR's repo to fetch the issue
    silently pulls whatever issue happens to share that number in the wrong
    repo. Fall back to the PR's repo only when the URL is missing/unparseable.
    """
    if url:
        match = _ISSUE_URL_RE.search(str(url))
        if match:
            return match.group(1)
    return fallback_repo


def load_spreadsheet_issues(path: str, sheet: Optional[str] = None) -> list[dict]:
    """Read Issues_v1.xlsx and return rows in the same format as load_spreadsheet().

    Each row must have a 'Closing PR #' to be eval-able (provides base_commit +
    test suite). Rows without a closing PR are skipped with a warning.

    The returned rows use the standard pipeline column keys so fetch_all() can
    process them identically to PRs.xlsx rows.  The issue number is stored under
    COL_ISSUE_NUMBER (as (repo, number) pairs) so instance_builder can fetch it
    directly instead of discovering it via PR body keyword matching.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet else wb.active
    headers = [cell.value for cell in ws[1]]
    rows_by_pr: dict[tuple[str, int], dict] = {}
    skipped = 0
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, raw_row))
        repo = row.get(COL_REPO)
        issue_number = row.get(_COL_ISSUE_NUMBER)
        closing_pr = row.get(_COL_CLOSING_PR)
        if not repo or not issue_number:
            continue
        if not closing_pr:
            logger.warning(
                f"Skipping {repo}#{issue_number}: no closing PR — cannot build eval instance."
            )
            skipped += 1
            continue
        issue_repo = _issue_repo_from_url(row.get(COL_URL), str(repo))
        if issue_repo != repo:
            logger.info(
                f"  Issue #{issue_number} URL points to {issue_repo}, not the "
                f"PR's repo {repo} -- fetching from {issue_repo}."
            )
        # One closing PR can resolve multiple issue rows. Merge them before the
        # PR-keyed ingest cache is consulted so no issue body is overwritten.
        key = (str(repo), int(closing_pr))
        if key not in rows_by_pr:
            rows_by_pr[key] = {
                COL_REPO: repo,
                COL_PR_NUMBER: int(closing_pr),
                COL_HAS_ISSUE: "Yes",
                _COL_ISSUE_NUMBER: [],
                COL_TITLE: [],
                COL_CATEGORY: row.get("Type", ""),
                COL_ALGORITHM_NAME: "",
                COL_PAPER_REFERENCE: "",
                COL_HAS_TEST: "Yes",
                COL_TEST_LINKS: "",
            }
        grouped = rows_by_pr[key]
        grouped[_COL_ISSUE_NUMBER].append((issue_repo, int(issue_number)))
        if row.get("Title"):
            grouped[COL_TITLE].append(str(row["Title"]))
    rows = list(rows_by_pr.values())
    for row in rows:
        row[COL_TITLE] = " | ".join(row[COL_TITLE])
    if skipped:
        logger.info(f"Skipped {skipped} issue row(s) with no closing PR.")
    logger.info(f"Loaded {len(rows)} PR group(s) with paired issues from {path}")
    return rows


def load_spreadsheet_rows(path: str, sheet: Optional[str] = None) -> list[dict]:
    """Load either the PR-oriented or issue-oriented spreadsheet format."""
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        ws = wb[sheet] if sheet else wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    finally:
        wb.close()
    if _COL_ISSUE_NUMBER in headers and COL_PR_NUMBER not in headers:
        return load_spreadsheet_issues(path, sheet=sheet)
    return load_spreadsheet(path, sheet=sheet)


def fetch_pr_data(repo: Repo, pr_number: int) -> Optional[dict]:
    """Fetch PR metadata from GitHub."""
    pull = repo.call_api(
        repo.api.pulls.get,
        owner=repo.owner,
        repo=repo.name,
        pull_number=pr_number,
    )
    if pull is None:
        logger.warning(f"PR {repo.owner}/{repo.name}#{pr_number} not found")
        return None
    return pull


def fetch_issue_data(repo: Repo, issue_number: int) -> Optional[dict]:
    """Fetch issue metadata from GitHub."""
    issue = repo.call_api(
        repo.api.issues.get,
        owner=repo.owner,
        repo=repo.name,
        issue_number=int(issue_number),
    )
    return issue


def find_linked_issue_numbers(pull) -> list[str]:
    """
    Extract issue numbers referenced by a PR using keyword matching.
    Replicates logic from swebench/collect/utils.py::Repo.extract_resolved_issues
    but works without a Repo instance (no commit fetching).
    """
    issues_pat = re.compile(r"(\w+)\s+\#(\d+)")
    comments_pat = re.compile(r"(?s)<!--.*?-->")
    text = (pull.title or "") + "\n" + (pull.body or "")
    text = comments_pat.sub("", text)
    references = issues_pat.findall(text)
    resolved = set()
    for word, issue_num in references:
        if word.lower() in PR_KEYWORDS:
            resolved.add(issue_num)
    return list(resolved)


def _row_cache_key(repo_full: str, pr_number: int) -> str:
    return f"{repo_full}#{pr_number}"


def _load_ingest_cache(cache_path: Path) -> dict[str, dict]:
    """Load previously-fetched rows from the ingest cache. Returns {key: row}."""
    if not cache_path.exists():
        return {}
    cache: dict[str, dict] = {}
    with open(cache_path) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
                key = _row_cache_key(row[COL_REPO], row[COL_PR_NUMBER])
                cache[key] = row
            except Exception:
                pass
    logger.info(f"Ingest cache: loaded {len(cache)} previously-fetched rows from {cache_path}")
    return cache


def _append_ingest_cache(cache_path: Path, row: dict) -> None:
    """Append a single enriched row to the ingest cache file."""
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    # pr_data and issue_data may contain GitHub API objects; serialize to plain dicts.
    serializable = {}
    for k, v in row.items():
        if k == "pr_data" and v is not None and not isinstance(v, dict):
            serializable[k] = dict(v) if hasattr(v, "__iter__") else str(v)
        elif k == "issue_data" and isinstance(v, dict):
            serializable[k] = {inum: (dict(iss) if not isinstance(iss, dict) else iss)
                                for inum, iss in v.items()}
        else:
            serializable[k] = v
    with open(cache_path, "a") as f:
        f.write(json.dumps(serializable, default=str) + "\n")


def fetch_all(
    spreadsheet_path: str,
    github_token: Optional[str] = None,
    limit: Optional[int] = None,
    pr_numbers: Optional[dict[str, set[int]]] = None,
    repos: Optional[set[str]] = None,
    issue_types: Optional[set[str]] = None,
    cache_path: Optional[Path] = None,
    sheet: Optional[str] = None,
) -> list[dict]:
    """
    Parse the spreadsheet and fetch GitHub data for each PR.

    Args:
        pr_numbers: optional {repo_full_name: {pr_number, ...}} filter.
                    When given, only those specific PRs are fetched.
        cache_path: if given, previously-fetched rows are loaded from this JSONL
                    file and skipped; newly fetched rows are appended immediately
                    so a crash mid-run loses at most the current row.

    Returns a list of enriched row dicts with added keys:
      - pr_data: raw GitHub PR object (or None)
      - issue_numbers: list of linked issue number strings
      - issue_data: dict mapping issue_number -> GitHub issue object
    """
    if not github_token:
        raise ValueError(
            "A GitHub token is required to fetch PR/issue data.\n"
            "Pass --github_token YOUR_TOKEN or set the GITHUB_TOKEN environment variable.\n"
            "Create one at: https://github.com/settings/tokens (no scopes needed for public repos)."
        )
    rows = load_spreadsheet_rows(spreadsheet_path, sheet=sheet)

    # Filter rows before hitting the GitHub API
    if repos:
        rows = [r for r in rows if r[COL_REPO] in repos]
        logger.info(f"Filtered spreadsheet to {len(rows)} row(s) matching --repos")
    if issue_types:
        rows = [r for r in rows if normalize_issue_type(r.get(COL_CATEGORY)) in issue_types]
        logger.info(
            f"Filtered spreadsheet to {len(rows)} row(s) matching --issue_types="
            f"{sorted(issue_types)}"
        )
    if pr_numbers:
        rows = [
            r for r in rows
            if r[COL_REPO] in pr_numbers and r[COL_PR_NUMBER] in pr_numbers[r[COL_REPO]]
        ]
        logger.info(f"Filtered spreadsheet to {len(rows)} row(s) matching --instance_ids")

    if limit:
        rows = rows[:limit]

    # Load checkpoint cache so we can skip already-fetched rows
    existing: dict[str, dict] = _load_ingest_cache(cache_path) if cache_path else {}

    # Cache Repo objects per owner/name
    repo_cache: dict[str, Repo] = {}

    enriched = []
    skipped = 0
    for i, row in enumerate(rows):
        repo_full = row[COL_REPO]  # e.g. "numpy/numpy"
        pr_number = row[COL_PR_NUMBER]
        owner, name = repo_full.split("/", 1)
        key = _row_cache_key(repo_full, pr_number)

        # Resume: use cached row if available
        if key in existing:
            enriched.append(existing[key])
            skipped += 1
            continue

        logger.info(f"[{i+1}/{len(rows)}] Fetching {repo_full}#{pr_number}")

        if repo_full not in repo_cache:
            try:
                repo_cache[repo_full] = Repo(owner, name, token=github_token)
            except Exception as e:
                logger.error(f"Failed to init Repo for {repo_full}: {e}")
                row["pr_data"] = None
                row["issue_numbers"] = []
                row["issue_data"] = {}
                enriched.append(row)
                if cache_path:
                    _append_ingest_cache(cache_path, row)
                continue

        repo = repo_cache[repo_full]
        pull = fetch_pr_data(repo, pr_number)
        row["pr_data"] = pull

        if pull is None:
            row["issue_numbers"] = []
            row["issue_data"] = {}
            enriched.append(row)
            if cache_path:
                _append_ingest_cache(cache_path, row)
            continue

        has_issue_flag = str(row.get(COL_HAS_ISSUE) or "").strip().lower()
        if has_issue_flag == "no":
            row["issue_numbers"] = []
            row["issue_data"] = {}
            logger.debug(f"  Skipping issue scan for {repo_full}#{pr_number} (Has Issue=No)")
            enriched.append(row)
            if cache_path:
                _append_ingest_cache(cache_path, row)
            continue

        # If the row already has a direct issue number (Issues_v1.xlsx), use it;
        # otherwise mine it from the PR body via keyword matching. Direct
        # entries may be bare numbers (issue lives in the PR's own repo) or
        # (issue_repo, number) pairs (issue lives in a different repo, as
        # recorded by load_spreadsheet_issues from the row's URL column).
        direct_issue = row.get(_COL_ISSUE_NUMBER)
        if direct_issue is not None:
            direct_issues = (
                direct_issue
                if isinstance(direct_issue, (list, tuple, set))
                else [direct_issue]
            )
            issue_refs = [
                entry
                if isinstance(entry, (list, tuple))
                else (repo_full, int(entry))
                for entry in direct_issues
            ]
            logger.debug(
                f"  Using direct issue numbers {issue_refs} for "
                f"{repo_full}#{pr_number}"
            )
        else:
            issue_refs = [(repo_full, int(n)) for n in find_linked_issue_numbers(pull)]
        issue_numbers = [str(inum) for _, inum in issue_refs]
        row["issue_numbers"] = issue_numbers

        issue_data = {}
        for issue_repo, inum in issue_refs:
            if issue_repo == repo_full:
                issue_repo_obj = repo
            else:
                if issue_repo not in repo_cache:
                    owner_i, name_i = issue_repo.split("/", 1)
                    try:
                        repo_cache[issue_repo] = Repo(owner_i, name_i, token=github_token)
                    except Exception as e:
                        logger.error(f"Failed to init Repo for {issue_repo}: {e}")
                        continue
                issue_repo_obj = repo_cache[issue_repo]
            issue = fetch_issue_data(issue_repo_obj, inum)
            if issue is not None:
                issue_data[str(inum)] = issue
        row["issue_data"] = issue_data

        if not issue_numbers:
            logger.warning(f"  No linked issues found for {repo_full}#{pr_number}")

        enriched.append(row)
        if cache_path:
            _append_ingest_cache(cache_path, row)

    if skipped:
        logger.info(f"Resumed from cache: skipped {skipped}/{len(rows)} already-fetched rows")
    return enriched
