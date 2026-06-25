"""
Issue-only inference pipeline.

Reads Issues_v1.xlsx (or any compatible sheet), fetches the issue body from GitHub,
identifies relevant repo files, and calls an LLM to produce a code patch.

No PR, no gold patch, no eval — pure inference.

Usage:
    python -m swebench.issue_pipeline.run_issue_pipeline \
        --spreadsheet Issues_v1.xlsx \
        --model deepseek-v4-flash \
        --endpoint https://api.deepseek.com/v1 \
        --github_token $GITHUB_TOKEN \
        --run_id openmm_issues_001
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Optional

import openpyxl
import requests
from tqdm.auto import tqdm

# Reuse inference helpers from the existing pipeline
from swebench.eval_pipeline.inference import (
    _extract_diff,
    _clean_patch,
    _repair_patch,
    _load_existing_ids,
    make_clients,
    _call_model,
    _calc_cost,
)
from swebench.eval_pipeline.constants import PATCH_INSTRUCTION

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# ── spreadsheet columns ───────────────────────────────────────────────────────
COL_REPO = "Repo"
COL_ISSUE_NUMBER = "Issue Number"
COL_TITLE = "Title"
COL_URL = "URL"
COL_TYPE = "Type"

# ── GitHub helpers ────────────────────────────────────────────────────────────

_GH_API = "https://api.github.com"
_SESSION_LOCK = threading.Lock()
_session: Optional[requests.Session] = None


def _gh_session(token: str) -> requests.Session:
    global _session
    if _session is None:
        s = requests.Session()
        s.headers.update({
            "Authorization": f"Bearer {token}",
            "Accept": "application/vnd.github+json",
            "X-GitHub-Api-Version": "2022-11-28",
        })
        with _SESSION_LOCK:
            if _session is None:
                _session = s
    return _session


def fetch_issue(repo: str, issue_number: int, token: str) -> dict:
    """Return GitHub issue JSON (title + body + comments)."""
    s = _gh_session(token)
    owner, name = repo.split("/", 1)
    r = s.get(f"{_GH_API}/repos/{owner}/{name}/issues/{issue_number}")
    r.raise_for_status()
    data = r.json()

    # Fetch comments too so the LLM has full context
    comments_url = data.get("comments_url", "")
    comments = []
    if comments_url and data.get("comments", 0) > 0:
        cr = s.get(comments_url)
        if cr.ok:
            for c in cr.json():
                body = (c.get("body") or "").strip()
                user = c.get("user", {}).get("login", "user")
                if body:
                    comments.append(f"**{user}**: {body}")

    return {
        "title": data.get("title", ""),
        "body": data.get("body", "") or "",
        "comments": comments,
        "state": data.get("state", ""),
        "labels": [l["name"] for l in data.get("labels", [])],
    }


def search_relevant_files(repo: str, issue_title: str, issue_body: str, token: str,
                          max_files: int = 10) -> list[dict]:
    """
    Use GitHub code search to find repo files mentioned in or relevant to the issue.
    Falls back to keyword extraction from the issue text if search API quota is hit.

    Returns list of {path, content} dicts.
    """
    s = _gh_session(token)
    owner, name = repo.split("/", 1)
    full_text = issue_title + " " + issue_body

    # Extract candidate identifiers: CamelCase, snake_case, file paths, function names
    token_pat = re.compile(r"\b([A-Za-z][A-Za-z0-9_]{3,})\b")
    raw_tokens = token_pat.findall(full_text)
    # Deduplicate, prefer longer tokens (more specific), skip common English words
    stop = {
        "this", "that", "with", "from", "have", "been", "will", "would", "could",
        "should", "their", "there", "when", "where", "which", "what", "about",
        "also", "some", "more", "than", "into", "your", "https", "http",
    }
    candidates = list(dict.fromkeys(
        t for t in raw_tokens if t.lower() not in stop and len(t) >= 4
    ))[:30]

    # Also look for explicit file paths
    path_pat = re.compile(r"[\w/]+\.(?:py|cpp|h|cu|cuh|java|f90|f)\b")
    explicit_paths = path_pat.findall(full_text)

    found_files: dict[str, str] = {}  # path → content

    def _fetch_file(path: str) -> Optional[tuple[str, str]]:
        """Fetch raw file content from GitHub, return (path, content) or None."""
        r = s.get(f"{_GH_API}/repos/{owner}/{name}/contents/{path}")
        if not r.ok:
            return None
        data = r.json()
        if isinstance(data, list):
            return None
        download_url = data.get("download_url")
        if not download_url:
            return None
        raw = s.get(download_url)
        if not raw.ok:
            return None
        return path, raw.text

    # Try explicit paths first (highest confidence)
    for path in explicit_paths:
        if len(found_files) >= max_files:
            break
        result = _fetch_file(path.lstrip("/"))
        if result:
            found_files[result[0]] = result[1]

    # GitHub code search for key identifiers
    for kw in candidates:
        if len(found_files) >= max_files:
            break
        try:
            time.sleep(1.2)  # GitHub search API: max 10 req/min for authenticated users
            resp = s.get(f"{_GH_API}/search/code", params={
                "q": f"{kw} repo:{repo}",
                "per_page": 3,
            })
            if resp.status_code in (403, 429):
                logger.warning("GitHub code search quota exceeded; stopping search")
                break
            if not resp.ok:
                continue
            items = resp.json().get("items", [])
            for item in items:
                if len(found_files) >= max_files:
                    break
                path = item.get("path", "")
                if path and path not in found_files:
                    result = _fetch_file(path)
                    if result:
                        found_files[result[0]] = result[1]
        except Exception as e:
            logger.debug(f"Search error for {kw!r}: {e}")
            continue

    logger.info(f"  Found {len(found_files)} relevant files for issue")
    return [{"path": p, "content": c} for p, c in found_files.items()]


# ── prompt builder ────────────────────────────────────────────────────────────

_MAX_FILE_CHARS = 200_000
_SYSTEM = (
    "You are an expert software engineer. "
    "You will be given a GitHub issue and must produce a git patch to resolve it."
)


def _format_files(files: list[dict]) -> str:
    if not files:
        return ""
    parts = [
        "Here are the current contents of the files most relevant to this issue. "
        "Each line is prefixed with its 1-based line number followed by a tab — "
        "use these EXACT line numbers when constructing your @@ hunk headers.\n"
    ]
    for f in files:
        content = f["content"]
        if len(content) > _MAX_FILE_CHARS:
            content = content[:_MAX_FILE_CHARS] + "\n... [truncated]"
        numbered = "\n".join(
            f"{i}\t{line}" for i, line in enumerate(content.split("\n"), start=1)
        )
        parts.append(f'<file path="{f["path"]}">\n{numbered}\n</file>')
    return "\n".join(parts) + "\n\n"


def build_prompt(issue: dict, files: list[dict], repo: str) -> str:
    title = issue["title"].strip()
    body = issue["body"].strip()
    comments = issue.get("comments", [])

    issue_text = f"**{title}**\n\n{body}"
    if comments:
        issue_text += "\n\n---\n**Discussion:**\n" + "\n\n".join(comments[:5])

    file_ctx = _format_files(files)

    return (
        f"{_SYSTEM}\n"
        f"Repository: {repo}\n\n"
        f"Here is the GitHub issue that needs to be resolved:\n"
        f"<issue>\n{issue_text}\n</issue>\n\n"
        f"{file_ctx}"
        f"{PATCH_INSTRUCTION}"
    )


# ── spreadsheet loader ────────────────────────────────────────────────────────

def load_issues(path: str, sheet: Optional[str] = None) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet else wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, raw))
        if not row.get(COL_REPO) or not row.get(COL_ISSUE_NUMBER):
            continue
        row[COL_ISSUE_NUMBER] = int(row[COL_ISSUE_NUMBER])
        # Build a stable instance_id
        repo_slug = row[COL_REPO].replace("/", "__")
        row["instance_id"] = f"{repo_slug}-{row[COL_ISSUE_NUMBER]}"
        rows.append(row)
    return rows


# ── cache helpers ─────────────────────────────────────────────────────────────

def _load_instance_cache(path: Path) -> dict[str, dict]:
    if not path.exists():
        return {}
    cache = {}
    with open(path) as f:
        for line in f:
            try:
                obj = json.loads(line)
                cache[obj["instance_id"]] = obj
            except Exception:
                pass
    return cache


def _save_instance(path: Path, obj: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a") as f:
        f.write(json.dumps(obj, default=str) + "\n")


# ── main pipeline ─────────────────────────────────────────────────────────────

def run_pipeline(args: argparse.Namespace) -> None:
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    instance_cache_path = out_dir / "instances.jsonl"
    predictions_path = out_dir / f"{args.model.replace('/', '_')}.{args.run_id}.jsonl"

    # Stage 1: load spreadsheet
    rows = load_issues(args.spreadsheet, sheet=args.sheet)
    if args.limit:
        rows = rows[: args.limit]
    if args.instance_ids:
        ids = set(args.instance_ids)
        rows = [r for r in rows if r["instance_id"] in ids]
    logger.info(f"Loaded {len(rows)} issues from {args.spreadsheet}")

    if not args.skip_fetch:
        # Stage 2: fetch issue bodies + relevant files
        instance_cache = _load_instance_cache(instance_cache_path)
        instances = []
        for row in rows:
            iid = row["instance_id"]
            if iid in instance_cache:
                instances.append(instance_cache[iid])
                continue

            repo = row[COL_REPO]
            issue_number = row[COL_ISSUE_NUMBER]
            logger.info(f"Fetching {repo}#{issue_number}")
            try:
                issue = fetch_issue(repo, issue_number, args.github_token)
                files = search_relevant_files(
                    repo, issue["title"], issue["body"], args.github_token,
                    max_files=args.max_files,
                )
                obj = {
                    "instance_id": iid,
                    "repo": repo,
                    "issue_number": issue_number,
                    "issue_title": issue["title"],
                    "issue_body": issue["body"],
                    "issue_comments": issue.get("comments", []),
                    "files": files,
                    "spreadsheet_row": {k: v for k, v in row.items() if k not in ("instance_id",)},
                }
                _save_instance(instance_cache_path, obj)
                instances.append(obj)
            except Exception as e:
                logger.error(f"Failed to fetch {iid}: {e}")
                obj = {
                    "instance_id": iid,
                    "repo": repo,
                    "issue_number": issue_number,
                    "error": str(e),
                    "files": [],
                }
                instances.append(obj)
    else:
        # Load from cache
        instance_cache = _load_instance_cache(instance_cache_path)
        instances = [instance_cache[r["instance_id"]] for r in rows if r["instance_id"] in instance_cache]
        logger.info(f"Loaded {len(instances)} instances from cache (--skip_fetch)")

    # Stage 3: build prompts
    prompts: dict[str, str] = {}
    for inst in instances:
        if inst.get("error"):
            continue
        issue = {
            "title": inst.get("issue_title", ""),
            "body": inst.get("issue_body", ""),
            "comments": inst.get("issue_comments", []),
        }
        prompts[inst["instance_id"]] = build_prompt(issue, inst.get("files", []), inst["repo"])

    logger.info(f"Built prompts for {len(prompts)}/{len(instances)} instances")

    if args.skip_inference:
        logger.info("--skip_inference set, stopping after prompt build")
        return

    # Stage 4: inference
    anthropic_client, openai_compat_client = make_clients(
        args.model, endpoint=args.endpoint, api_key=args.api_key
    )

    existing_ids = _load_existing_ids(str(predictions_path), model_name=args.model)
    todo = [i for i in instances if i["instance_id"] not in existing_ids]
    logger.info(f"Running inference on {len(todo)} instances (skipping {len(existing_ids)} already done)")

    total_cost = 0.0
    cost_lock = threading.Lock()
    write_lock = threading.Lock()

    def _process(inst: dict) -> None:
        nonlocal total_cost
        iid = inst["instance_id"]
        prompt = prompts.get(iid)
        if not prompt:
            record = {"instance_id": iid, "model_patch": "", "model_name_or_path": args.model, "skipped": True}
            with write_lock:
                with open(predictions_path, "a") as f:
                    f.write(json.dumps(record) + "\n")
            return

        try:
            response_text, cost = _call_model(
                prompt, args.model,
                anthropic_client=anthropic_client,
                openai_compat_client=openai_compat_client,
                max_tokens=args.max_tokens,
            )
            patch = _extract_diff(response_text)
            patch = _clean_patch(patch)
            patch = _repair_patch(patch)

            with cost_lock:
                total_cost += cost
            logger.info(f"{iid}: cost=${cost:.4f}, total=${total_cost:.2f}")

            record = {
                "instance_id": iid,
                "model_patch": patch,
                "model_name_or_path": args.model,
                "full_output": response_text,
            }
        except Exception as e:
            logger.error(f"Error on {iid}: {e}")
            record = {"instance_id": iid, "model_patch": "", "model_name_or_path": args.model, "error": str(e)}

        with write_lock:
            with open(predictions_path, "a") as f:
                f.write(json.dumps(record) + "\n")

    with ThreadPoolExecutor(max_workers=args.max_workers) as pool:
        futs = {pool.submit(_process, inst): inst for inst in todo}
        with tqdm(total=len(todo), desc="Inference") as pbar:
            for fut in as_completed(futs):
                fut.result()
                pbar.update(1)

    logger.info(f"Done. Predictions written to {predictions_path}")
    logger.info(f"Total cost: ${total_cost:.4f}")

    # Stage 5: print summary
    results = []
    with open(predictions_path) as f:
        for line in f:
            try:
                results.append(json.loads(line))
            except Exception:
                pass

    non_empty = sum(1 for r in results if r.get("model_patch", "").strip())
    skipped = sum(1 for r in results if r.get("skipped"))
    errors = sum(1 for r in results if r.get("error"))

    print("\n" + "=" * 70)
    print(f"  ISSUE PIPELINE RESULTS  ({args.run_id})")
    print("=" * 70)
    print(f"  Model:          {args.model}")
    print(f"  Total issues:   {len(results)}")
    print(f"  With patches:   {non_empty}")
    print(f"  Skipped:        {skipped}")
    print(f"  Errors:         {errors}")
    print(f"  Total cost:     ${total_cost:.4f}")
    print(f"  Output:         {predictions_path}")
    print("=" * 70)


# ── CLI ───────────────────────────────────────────────────────────────────────

def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Issue-only LLM inference pipeline")
    p.add_argument("--spreadsheet", default="Issues_v1.xlsx", help="Path to issues spreadsheet")
    p.add_argument("--sheet", default=None, help="Sheet name (default: active sheet)")
    p.add_argument("--model", required=True, help="Model name")
    p.add_argument("--endpoint", default=None, help="OpenAI-compatible API endpoint URL")
    p.add_argument("--api_key", default=None, help="API key (defaults to env vars)")
    p.add_argument("--github_token", default=os.environ.get("GITHUB_TOKEN"), help="GitHub token")
    p.add_argument("--run_id", default="issues_run", help="Run identifier for output files")
    p.add_argument("--output_dir", default=None, help="Output directory (default: outputs/<run_id>)")
    p.add_argument("--max_tokens", type=int, default=32768)
    p.add_argument("--max_workers", type=int, default=4)
    p.add_argument("--max_files", type=int, default=10, help="Max relevant files to fetch per issue")
    p.add_argument("--limit", type=int, default=None, help="Limit number of issues")
    p.add_argument("--instance_ids", nargs="*", default=None, help="Filter to specific instance IDs")
    p.add_argument("--skip_fetch", action="store_true", help="Skip GitHub fetch, use cached instances")
    p.add_argument("--skip_inference", action="store_true", help="Stop after building prompts")
    return p.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    if args.output_dir is None:
        args.output_dir = f"outputs/{args.run_id}"
    if not args.github_token and not args.skip_fetch:
        print("ERROR: --github_token required (or set GITHUB_TOKEN env var)", file=sys.stderr)
        sys.exit(1)
    run_pipeline(args)
